from openpyxl import load_workbook, Workbook
from scipy import stats
import math
import xlsxwriter

class ShoreLineAngles:
    """
    This Class takes angles of the shore profile from a piece of shore.
    Export the shore angles from ArcMap as a table and as an excel file.
    It is programmed to work with EXCEL FILES ONLY
    """

    def __init__(self, file_name):
        self.wb = load_workbook('./PreRunFiles/ShoreProfiles' + file_name + '.xlsx') # The workbook from the excel sheet
        self.ws = self.wb.active  # The worksheet from the workbook
        self.regression_angles = list() # list of regression angles for each profile
        self.average_angles = list() # list of average angles for each profile
        self.ROW_OF_FIRST_POINT = 2
        return


    def getAveragesSlopeAngles(self, profiles_to_average):
        """
        Uses the coordinates of the first and last points in the shoreline profile
        to calculate the angles of the average slope for each shoreline profile
        :return:  Array of average angles for each shoreline profile in order.
                    (Not necessary as it is saved as a member variable)
        """


        for i in range(1, self.ws.max_column + 1, 2): #Loops through each profile as they take up two columns each

            # Get the first point values for current profile
            min_col_1 = self.ws.iter_cols(min_col=i, min_row=self.ROW_OF_FIRST_POINT, max_col=i, max_row=self.ROW_OF_FIRST_POINT)
            min_col_2 = self.ws.iter_cols(min_col=i+1, min_row=self.ROW_OF_FIRST_POINT, max_col=i+1, max_row=self.ROW_OF_FIRST_POINT)

            for col in min_col_1:
                for cell in col:
                    x1 = cell.value
            for col in min_col_2:
                for cell in col:
                    y1 = cell.value

            # Get the last point values for current profile
            max_rows = len(self.ws[xlsxwriter.utility.xl_col_to_name(i-1)])
            max_col_1 = self.ws.iter_cols(min_col=i, min_row=self.ROW_OF_FIRST_POINT, max_col=i, max_row=self.ws.max_row)
            max_col_2 = self.ws.iter_cols(min_col=i+1, min_row=self.ROW_OF_FIRST_POINT, max_col=i+1, max_row=self.ws.max_row)


            for col in max_col_1:
                for cell in col:
                    if (cell.value is not None):
                        x2 = cell.value
            for col in max_col_2:
                for cell in col:
                    if (cell.value is not None):
                        y2 = cell.value

            # Average out the slope and take the inverse tan to find the angle
            # print(y2, y1, x2, x1)
            average_slope = (y2 - y1)/(x2 - x1)
            self.average_angles.append(math.degrees(math.atan(average_slope)))

        # print(self.average_angles)
        average_angle = 0
        for profile in profiles_to_average:
            average_angle += self.average_angles[profile - 1]

        average_angle /= len(profiles_to_average)

        return average_angle

    def getRegressionLineAngles(self, profiles_to_average):
        """
        Uses the coordinates of each point in each shoreline profile
        to calculate an angle for the regression line slope of each profile
        :return: Array of regression angles for each profile in order.
                    (Not necessary as it is saved as a member variable)
        """
        for i in range(1, self.ws.max_column + 1, 2): # Loops through each profile as they take up two columns each
            xCord = []
            yCord = []
            cols = self.ws.iter_cols(min_col=i, min_row=0, max_col=i+1, max_row=self.ws.max_row)
            for col in cols:
                is_x_cord = False
                for cell in col:
                    if(cell.value == 'X'): # Lets the program know if it is the x column or the y column for the current profile
                        is_x_cord = True
                        continue
                    if(is_x_cord):
                        if(cell.value is not None):
                            xCord.append(cell.value)
                    else:
                        if(cell.value is not None):
                            yCord.append(cell.value)

            yCord = yCord[1:]

            # Get the slope of the regression line and take the inverse tan to find the angle value
            slope, intercept, r_value, p_value, std_err = stats.linregress(xCord, yCord)
            self.regression_angles.append(math.degrees(math.atan(slope)))

        # print(self.regression_angles)
        average_angle = 0
        for profile in profiles_to_average:
            average_angle += self.average_angles[profile - 1]

        average_angle /= len(profiles_to_average)

        return average_angle

    def generateExcelSheet(self):
        """
        Puts the average and regression angles calculated into another sheet in the opened
        workbook the initial date came from
        :return: None
        """

        angles_ws = self.wb.create_sheet('angles')

        angles_ws.append(['Profile Number'] + [i + 1 for i in range(len(self.regression_angles))])
        angles_ws.append(['Latitude'] + [i + 1 for i in range(len(self.regression_angles))])
        angles_ws.append(['Longitude'] + [i + 1 for i in range(len(self.regression_angles))])
        angles_ws.append(['Regression Slope Angle'] + [angle for angle in self.regression_angles])
        angles_ws.append(['Average Slope Angle'] + [angle for angle in self.average_angles])

        self.wb.save('Slope_File2.xlsx')

        return
